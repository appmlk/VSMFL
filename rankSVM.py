# -*- coding: utf-8 -*-
import openpyxl
from sklearn import svm, linear_model, model_selection
from sklearn.metrics import r2_score
import numpy as np
import os
from matplotlib import pyplot as plt 
import Statistics
# from pysofia.compat import RankSVM

def cal_correlation(file_path1, file_path2):
    '''
    计算两种技术的相关性
    '''
    tec1 = file_path1.split('\\')[-1].split('_')[0]
    tec2 = file_path2.split('\\')[-1].split('_')[0]
    wb1 = openpyxl.load_workbook(file_path1) # 读取 工作簿
    wb2 = openpyxl.load_workbook(file_path2) # 读取 工作簿
    x_values = []
    y_values = []
    for i, sheet1 in enumerate(wb1):
        if i == 0 or i == 1:
            continue
        print(sheet1.title)
        sheet2 = wb2.worksheets[i]  
        for j, row in enumerate(sheet1.rows):
            if j == 0:
                continue
            name = row[0].value
            final_sus1 = row[3].value
            final_sus2 = sheet2.cell(row = j + 1, column=4).value
            # print(name)
            x_values.append(final_sus1)
            y_values.append(final_sus2)
            # cal_correlation_single(final_sus1, final_sus2)
            # break
        # break
    # print(tec1, tec2)
    plt.title('The correlation between %s and %s' % (tec1, tec2)) 
    plt.xlabel("exam of %s" % tec1) 
    plt.ylabel("exam of %s" % tec2) 
    # print(np.array(x_values)  ,np.array(y_values))
    plt.scatter(np.array(x_values)  ,np.array(y_values)) 
    # plt.show()
    plt.savefig('%s_%s.png' % (tec1, tec2))
    plt.close()
    return

def cal_r2_score(file_path1, file_path2):
    '''
    计算相关系数r方
    '''
    tec1 = file_path1.split('\\')[-1].split('_')[0]
    tec2 = file_path2.split('\\')[-1].split('_')[0]
    wb1 = openpyxl.load_workbook(file_path1) # 读取 工作簿
    wb2 = openpyxl.load_workbook(file_path2) # 读取 工作簿
    x_values = []
    y_values = []
    for i, sheet1 in enumerate(wb1):
        if i == 0 or i == 1:
            continue
        # print(sheet1.title)
        sheet2 = wb2.worksheets[i]  
        for j, row in enumerate(sheet1.rows):
            if j == 0:
                continue
            name = row[0].value
            final_sus1 = row[3].value
            final_sus2 = sheet2.cell(row = j + 1, column=4).value
            x_values.append(final_sus1)
            y_values.append(final_sus2)
    coefficient_of_dermination = r2_score(x_values, y_values)
    print(tec1 + ' ' + tec2 + ' is ' + str(coefficient_of_dermination))
    # coefficient_of_dermination = r2_score(y_values, x_values)
    # print(tec2 + ' ' + tec1 + ' is ' + str(coefficient_of_dermination))
    return

def normalization(data):
    _range = np.max(data) - np.min(data)
    if _range == 0:
        _range = 1
    return (data - np.min(data)) / _range
 
def standardization(data):
    mu = np.mean(data, axis=0)
    sigma = np.std(data, axis=0)
    return (data - mu) / sigma

def prepare_data(tec_list, tag_root_path, is_nomalization=True):
    '''
    准备数据
    '''
    train_X = {} # 用来装训练数据(也包括测试集!),里面每个key代表一份代码,元素是若干个技术的numpy向量
    train_Y = {} # 用来装训练数据(也包括测试集!),里面每个key代表一份代码,元素是该代码对应的错误行numpy向量
    maxn = 0
    for tec_path in tec_list:     
        tec = tec_path.split('/')[-1].split('_')[0]
        wb = openpyxl.load_workbook(tec_path) # 读取 工作簿
        for i, sheet in enumerate(wb):
            if i == 0 or i == 1:
                continue
            # print(sheet.title)
            tag_dir = os.path.join(tag_root_path, sheet.title, 'Tag_c')
            tag_info = Statistics.get_tag_info(tag_dir)
            for j, row in enumerate(sheet.rows):
                if j == 0:
                    continue
                name = row[0].value
                # print(name)
                final_sus = eval(row[2].value)
                if tec == 'vsmfl':
                    final_sus = np.array(final_sus)
                    final_sus = np.max(final_sus) - final_sus # 这里是VSMFL时，需要倒置
                else:
                    final_sus.insert(0, 0) # 这里是SBFL时，多加一个第零行,怀疑度自然为0
                    final_sus = np.array(final_sus)
                if is_nomalization:
                    final_sus = normalization(final_sus)
                maxn = max(maxn, final_sus.shape[0]) 

                if name not in train_X:
                    train_X[name] = {}
                    train_Y[name] = {}
                    Y = np.zeros(final_sus.size)
                    for line in tag_info[name]:
                        Y[line] = 1
                    train_Y[name] = Y
                train_X[name][tec] = final_sus
                
                # X = np.c_[final_sus1, final_sus2]
                # print(X, Y)
                # break
            # break
    # for key in train_X:
    #     for tec in train_X[key]:
    #         X = train_X[key][tec]
    #         X = np.pad(X ,((0, maxn - X.shape[0])),'constant')
    #         train_X[key][tec] = X
    #     Y = train_Y[key]
    #     Y = np.pad(Y, (0, maxn - Y.shape[0]), 'constant')
    #     train_Y[key] = Y
    # print(train_X, train_Y)
    return train_X, train_Y

def vectorization(train_X, train_Y):
    '''
    将多种技术得到的怀疑度结果向量化
    注意这里传进来的train_X和train_Y均为列表
    '''
    # model = RankSVM(max_iter=10, alpha = 0.1)
    X_train_list = []
    Y_train_list = []
    IS_FIRST = True
    for i, id in enumerate(train_X):
        flag = True
        for tec in id:
            if flag:
                X = id[tec]
                flag = False
            else:
                X = np.c_[X, id[tec]]
        Y = train_Y[i]
        if IS_FIRST:
            X_train_list = X
            Y_train_list = Y
            IS_FIRST = False
        else:
            X_train_list = np.vstack((X_train_list, X))
            Y_train_list = np.append(Y_train_list, Y)
    # X_list = np.array(X_list)
    # Y_list = np.array(Y_list)
    
    return X_train_list, Y_train_list

def predict_result(RSVM, X_test_list, Y_test_list):
    '''
    预测结果
    '''
    file_path = r'result/ltr.xlsx'
    wb = openpyxl.load_workbook(file_path)
    problem_id = '2710'
    wb.create_sheet(problem_id)
    ws = wb[problem_id]
    # print(ws.max_row)
    if ws.max_row == 1:
        ws.append({'a':'name'})
    now_column = ws.max_column
    print(now_column)
    for key in X_test_list:
        # print(Y_test_list[key])
        res = RSVM.predict(X_test_list[key])
        print(res)
        if now_column == 1:
            ws.append({'a':key, 'b': str(res)})
        else:
            for i, row in enumerate(ws.rows):
                if row[0].value == key:
                    ws.cell(row=i, column=now_column+1, value=str(res))
                    break
    wb.save(file_path)
    return

if __name__ == "__main__":


    tec_list = ['Jaccard','Tarantula','Op2','Ochiai', 'Dstar', 'vsbfl']
    # tec_list = ['Jaccard', 'Dstar']
    for i, tec1 in enumerate(tec_list):
        for j, tec2 in enumerate(tec_list):
            if tec1 == 'vsbfl':
                file_path1 = os.path.join(r'result', 'vsbfl_sus.xlsx')
            else:
                file_path1 = os.path.join(r'result', 'sbfl(' + tec1 + ')_sus.xlsx')
            if tec2 == 'vsbfl':
                file_path2 = os.path.join(r'result', 'vsbfl_sus.xlsx')
            else:
                file_path2 = os.path.join(r'result', 'sbfl(' + tec2 + ')_sus.xlsx')
            # print( tec2)
            # cal_correlation(file_path1, file_path2)
            cal_r2_score(file_path1, file_path2)
            # break
        # break
    # file_path1 = r'result/vsbfl_sus.xlsx'
    # file_path2 = r'result/sbfl(Jaccard)_sus.xlsx'
    # tec_list = [file_path1, file_path2]
    # train_X, train_Y = prepare_data(tec_list, r'ITSP-data')
    # RSVM = learn_to_rank(train_X, train_Y)

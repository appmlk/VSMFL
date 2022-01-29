import rankSVM
import openpyxl
import numpy as np
import os
import Statistics

file_path = r'result/multiplication_fit(N)_sus.xlsx'
tag_root_dir = r'D:\fault_loc\ITSP-data'

def statistical_fl_results(file_path, tag_root_dir):
    '''
    统计错误定位的实验结果
    '''
    if file_path.split('.')[-1] != 'xlsx' and file_path.split('.')[-1] != 'xls':
        print('not an execl file')
        return
    wb = openpyxl.load_workbook(file_path) # 读取 工作簿
    ws_first = wb.worksheets[0]  
    ws_first.title = 'first'
    ws_multi = wb.worksheets[1]
    ws_multi.title = 'multi'
    ws_first.append({'a':'problem_id', 'b':'exam', 'c':'top_1', 'd':'top_3', 'e':'top_5', 'f': 'top_10','g': 'total_num'})
    ws_multi.append({'a':'problem_id', 'b':'exam', 'c':'top_1', 'd':'top_3', 'e':'top_5', 'f': 'top_10','g': 'total_num'})

    for i, sheet in enumerate(wb):
        if i == 0 or i == 1:
            continue
        ans1 = 0
        ans2 = 0
        # if sheet.title != '2812':
        #     continue
        print(sheet.title)
        # row_max = sheet.max_row # 获取最大行
        # con_max = sheet.max_column # 获取最大列
        tag_info = Statistics.get_all_tag_info(tag_root_dir)
        # for i in tag_info:
        #     ans1 += len(tag_info[i])
        # continue
        ans1 += Statistics.get_code_exam(sheet, tag_info) # 给每道题添加exam
        ans2 += (sheet.max_row - 1)
        exam = Statistics.cal_exam_first(sheet.rows, tag_info)
        top_1, top_3, top_5, top_10 = Statistics.cal_top_N_first(sheet.rows, tag_info)
        ws_first.append({'a':sheet.title, 'b':np.mean(exam), 'c':top_1, 'd':top_3, 'e':top_5, 'f': top_10, 'g':sheet.max_row - 1})
        
        top_1, top_3, top_5, top_10 = Statistics.cal_top_N_multi(sheet.rows, tag_info)
        exam = Statistics.cal_exam_multi(sheet.rows, tag_info)
        ws_multi.append({'a':sheet.title, 'b':np.mean(exam), 'c':top_1, 'd':top_3, 'e':top_5, 'f': top_10, 'g':sheet.max_row - 1})
        
        # print(exam)
        # print(top_1, top_3, top_5, top_10)
        # break
        print(ans1, ans2, ans1/ans2)
    wb.save(file_path)
    return


def start_fit(train_X, train_Y):
    '''
    将最终怀疑度定义为 alpha * 技术1 + (1 - alpha) * 技术2
    alpha从0.1取到0.9，间隔为0.1，计算实验结果
    '''
    alpha = 0
    while alpha <= 1:
        print(alpha)
        wb = openpyxl.load_workbook(file_path)
        sheet_id = str(alpha)
        wb.create_sheet(sheet_id)
        ws = wb[sheet_id]
        ws.append({'a':'name', 'b':'suspicion_rank', 'c':'suspicion'})
        for key in train_X:
            tecs = list(train_X[key].keys())
            sus1 = train_X[key][tecs[0]]
            sus2 = train_X[key][tecs[1]]
            final_sus = alpha * sus1 + (1-alpha) * sus2
            # print(sus1)
            # print(sus2)
            # print(final_sus)

            line_rank = []
            for i,sus in enumerate(final_sus):
                if i == 0:
                    continue
                line_rank.append({
                    'sus': sus,
                    'no': i
                })
            line_rank.sort(key=lambda s: s['sus'], reverse=True)
            # print(line_rank)
            final_rank = []
            tmp_list = []
            for i in range(len(line_rank)):
                if i != 0 and abs(line_rank[i]['sus'] - line_rank[i-1]['sus']) > 0.0001:
                    final_rank.append(tmp_list)
                    tmp_list = []
                tmp_list.append(line_rank[i]['no'])
            if len(tmp_list) != 0:
                final_rank.append(tmp_list)
            # print(final_rank)
            ws.append({'a': key, 'b': str(final_rank), 'c': str(final_sus)})
            # break
        wb.save(file_path)
        alpha += 0.1
        # break
    return

def line_fit(train_X, train_Y):
    '''
    将最终怀疑度定义为 (1 + 技术1) * (1 + 技术2)，计算实验结果
    '''
    wb = openpyxl.load_workbook(file_path)
    sheet_id = 'line'
    wb.create_sheet(sheet_id)
    ws = wb[sheet_id]
    ws.append({'a':'name', 'b':'suspicion_rank', 'c':'suspicion'})
    for key in train_X:
        print(key)
        tecs = list(train_X[key].keys())
        sus1 = train_X[key][tecs[0]]
        sus2 = train_X[key][tecs[1]]
        final_sus = (1 + sus1) * (1 + sus2)
        # print(sus1)
        # print(sus2)
        # print(final_sus)

        line_rank = []
        for i,sus in enumerate(final_sus):
            if i == 0:
                continue
            line_rank.append({
                'sus': sus,
                'no': i
            })
        line_rank.sort(key=lambda s: s['sus'], reverse=True)
        # print(line_rank)
        final_rank = []
        tmp_list = []
        for i in range(len(line_rank)):
            if i != 0 and abs(line_rank[i]['sus'] - line_rank[i-1]['sus']) > 0.0001:
                final_rank.append(tmp_list)
                tmp_list = []
            tmp_list.append(line_rank[i]['no'])
        if len(tmp_list) != 0:
            final_rank.append(tmp_list)
        # print(final_rank)
        ws.append({'a': key, 'b': str(final_rank), 'c': str(final_sus)})
        # break
    wb.save(file_path)
    return

if __name__ == "__main__":

    file_path1 = r'result/vsmfl_sus.xlsx'
    file_path2 = r'result/sbfl(Jaccard)_sus.xlsx'
    tec_list = [file_path1, file_path2]
    train_X, train_Y = rankSVM.prepare_data(tec_list, r'D:\fault_loc\ITSP-data', is_nomalization=False)
    # print(train_X)
    # start_fit(train_X, train_Y)
    line_fit(train_X, train_Y)

    statistical_fl_results(file_path, tag_root_dir)

from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures
from sklearn.pipeline import Pipeline
from sklearn.metrics import mean_squared_error
from sklearn import model_selection
import numpy as np
import rankSVM
import openpyxl
import Statistics

file_path = r'result/regression(N)_sus.xlsx'
tag_root_dir = r'D:\fault_loc\ITSP-data'

def statistical_fl_results(file_path, tag_root_dir):
    '''
    统计错误定位的实验结果
    '''
    if file_path.split('.')[-1] != 'xlsx' and file_path.split('.')[-1] != 'xls':
        print('not an execl file')
        return
    wb = openpyxl.load_workbook(file_path) # 读取 工作簿
    ws_first = wb.worksheets[0]  
    ws_first.title = 'first'
    ws_multi = wb.worksheets[1]
    ws_multi.title = 'multi'
    ws_first.append({'a':'problem_id', 'b':'exam', 'c':'top_1', 'd':'top_3', 'e':'top_5', 'f': 'top_10','g': 'total_num'})
    ws_multi.append({'a':'problem_id', 'b':'exam', 'c':'top_1', 'd':'top_3', 'e':'top_5', 'f': 'top_10','g': 'total_num'})
    ans1 = 0
    ans2 = 0
    for i, sheet in enumerate(wb):
        if i == 0 or i == 1:
            continue
        # if sheet.title != '2812':
        #     continue
        print(sheet.title)
        # row_max = sheet.max_row # 获取最大行
        # con_max = sheet.max_column # 获取最大列
        tag_info = Statistics.get_all_tag_info(tag_root_dir)
        # for i in tag_info:
        #     ans1 += len(tag_info[i])
        # continue
        ans1 += Statistics.get_code_exam(sheet, tag_info) # 给每道题添加exam
        ans2 += (sheet.max_row - 1)
        exam = Statistics.cal_exam_first(sheet.rows, tag_info)
        top_1, top_3, top_5, top_10 = Statistics.cal_top_N_first(sheet.rows, tag_info)
        ws_first.append({'a':sheet.title, 'b':np.mean(exam), 'c':top_1, 'd':top_3, 'e':top_5, 'f': top_10, 'g':sheet.max_row - 1})
        
        top_1, top_3, top_5, top_10 = Statistics.cal_top_N_multi(sheet.rows, tag_info)
        exam = Statistics.cal_exam_multi(sheet.rows, tag_info)
        ws_multi.append({'a':sheet.title, 'b':np.mean(exam), 'c':top_1, 'd':top_3, 'e':top_5, 'f': top_10, 'g':sheet.max_row - 1})
        
        # print(exam)
        # print(top_1, top_3, top_5, top_10)
        # break
        print(ans1, ans2, ans1/ans2)
    wb.save(file_path)
    return


def polynomial_model(degree=1):
    polynomial_features = PolynomialFeatures(degree=degree, include_bias=False)
    linear_regression = LinearRegression(normalize=True)
    pipeline = Pipeline([("polynomial_features", polynomial_features),
    ("linear_regression", linear_regression)])
    return pipeline

def start_regression(file_path, tec_list):
    '''
    '''
    wb = openpyxl.load_workbook(file_path)

    train_X, train_Y = rankSVM.prepare_data(tec_list, r'D:\fault_loc\ITSP-data', is_nomalization=False)
    solution_id_list = list(train_X.keys())
    degrees = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10]
    results = []
    for d in degrees:
        sheet_id = str(d)
        wb.create_sheet(sheet_id)
        ws = wb[sheet_id]
        ws.append({'a':'name', 'b':'suspicion_rank', 'c':'suspicion'})
        print(str(d))

        model = polynomial_model(degree=d)
        kf = model_selection.KFold(5, shuffle=True)
        for train_index, test_index in kf.split(solution_id_list):
            now_train_X = []
            now_train_Y = []
            for id in train_index:
                now_train_X.append(train_X[solution_id_list[id]])
                now_train_Y.append(train_Y[solution_id_list[id]])
            X_train_list, Y_train_list = rankSVM.vectorization(now_train_X, now_train_Y)
            print(X_train_list)

            model.fit(X_train_list, Y_train_list)
            # 相关测试
            # now_test_X = []
            # now_test_Y = []
            # for id in test_index:
            #     now_test_X.append(train_X[solution_id_list[id]])
            #     now_test_Y.append(train_Y[solution_id_list[id]])
            # X_test_list, Y_test_list = rankSVM.vectorization(now_test_X, now_test_Y)
            # train_score = model.score(X_test_list, Y_test_list)
            # mse = mean_squared_error(Y_test_list, model.predict(X_test_list))
            # results.append({"model": model, "degree": d, "score": train_score, "mse": mse})

            for i in test_index:
                solution_id = solution_id_list[i]
                test_X, test_Y = rankSVM.vectorization([train_X[solution_id]], [train_Y[solution_id]])
                sus = model.predict(test_X)
                # print(solution_id, sus)
                line_rank = []
                for i,sus in enumerate(sus):
                    if i == 0:
                        continue
                    line_rank.append({
                        'sus': sus,
                        'no': i
                    })
                line_rank.sort(key=lambda s: s['sus'], reverse=True)
                # print(line_rank)
                final_rank = []
                tmp_list = []
                for i in range(len(line_rank)):
                    if i != 0 and abs(line_rank[i]['sus'] - line_rank[i-1]['sus']) > 0.0001:
                        final_rank.append(tmp_list)
                        tmp_list = []
                    tmp_list.append(line_rank[i]['no'])
                if len(tmp_list) != 0:
                    final_rank.append(tmp_list)
                # print(final_rank)
                ws.append({'a': solution_id, 'b': str(final_rank), 'c': str(sus)})
            # break
        wb.save(file_path)
        # break
    # for r in results:
    #     print("degree: {}; train score: {}; mean squared error: {}".format(r["degree"], r["score"], r["mse"]))
    return 

if __name__ == "__main__":

    file_path1 = r'result/vsmfl_sus.xlsx'
    file_path2 = r'result/sbfl(Jaccard)_sus.xlsx'
    tec_list = [file_path1, file_path2]
    
    # start_regression(file_path, tec_list)
    statistical_fl_results(file_path, tag_root_dir)


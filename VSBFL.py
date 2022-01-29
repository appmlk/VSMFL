import Variable_sus as vs
import util
import os
import numpy as np
import openpyxl
import sys

alpha = 0.8 # 怀疑度系数
beta = 0.8  # 输出系数
data_path = r'D:\fault_loc\ITSP-data'
res_file = r''
res_file_temp = os.path.join('result', 'vsbfl_BUCT_sus.xlsx')  # 请先自己创建这个文件


def find_pair_by_tag(dir_path):
    '''
    根据数据集寻找每份错误代码对应的正确代码
    '''
    pair_info = {}
    tag_files = os.listdir(dir_path)
    for i in tag_files:
        tag_file = os.path.join(dir_path, i)
        lines = util.read_file(tag_file)
        wa_file = lines[0].replace('\n', '')
        ac_file = lines[1].replace('\n', '')
        pair_info[wa_file] = ac_file
    # for file in pair_info:
    #     print(file, pair_info[file])
    return pair_info

def find_pair_by_res(file_path):
    '''
    根据匹配结果寻找每份错误代码对应的正确代码
    '''
    pair_info = {}
    if file_path.split('.')[-1] != 'xlsx' and file_path.split('.')[-1] != 'xls':
        print('not an execl file')
        return
    wb = openpyxl.load_workbook(file_path)
    for i, sheet in enumerate(wb):
        if i == 0:
            continue
        # print(sheet.title)
        buggy_name = ''
        for j, row in enumerate(sheet.rows):
            if j == 0:
                continue
            if row[0].value != None:
                buggy_name = row[0].value
            if row[1].value.find('[') >= 0:
                ac_name_list = eval(row[1].value)
                pair_info[buggy_name] = ac_name_list[0]
    #     break
    # print(pair_info)
    return pair_info

def find_possiable_correct_line(wa_res, ac_res):
    '''
    寻找配对的变量
    '''
    # print(wa_res)
    weight, weight_LCS = vs.add_weight(wa_res, ac_res)
    vars_pair = util.cal_KM(weight) #二分图最大完备匹配
    # print(vars_pair)
    # print(weight_LCS)
    possiable_correct_line = {}
    for var in vars_pair:
        possiable_correct_line[var] = weight_LCS[var][vars_pair[var]['var']]
    # possiable_correct_line分为n个列表，n是测试用例的数量，每个列表中的数字为
    # 错误代码变量值序列中能和正确代码匹配的值在整个值序列中的位置
    # print(possiable_correct_line)
    return possiable_correct_line, vars_pair


def cal_sus(wa_file, ac_file, test_dir_path, language):
    '''
    计算怀疑度
    '''
    wa_res, ac_res = vs.get_sequences(wa_file, ac_file, test_dir_path, language)
    # print(wa_res)
    # print(ac_res)
    possiable_correct_line, vars_pair = find_possiable_correct_line(wa_res, ac_res)
    # print(possiable_correct_line)
    lines = util.read_file(wa_file)
    file_length = len(lines)
    line_sus = list(np.ones(file_length + 1))
    for var in possiable_correct_line:
        line_nums = possiable_correct_line[var]
        # print(var, line_nums)
        for i, test_case in enumerate(wa_res):
            # print(i, test_case['info'])
            if var not in test_case['info']:
                continue
            info = test_case['info'][var]
            line_num = line_nums[i]
            possiable_fault_line = range(0, len(info))
            possiable_fault_line = list(filter(lambda x: x not in line_num, possiable_fault_line))
            possiable_fault_line = list(map(lambda x: info[x]['line'], possiable_fault_line))
            line_num = list(map(lambda x: info[x]['line'], line_num))
            # print(var, i, line_num, len(ac_res[i]['info'][var]) - len(line_nums[i]))
            if test_case['res'] == False:
                for j in possiable_fault_line:
                    line_sus[int(j)] *= alpha
                # 这个补丁是为了解决计算不足的问题
                # 这里列表直接用-1还是有危险的，因为可能列表长度为0
                if(len(possiable_fault_line) == 0 and info[-1]['value'] != ac_res[i]['info'][vars_pair[var]['var']][-1]['value']):
                    print(var, i, info)
                    line_sus[int(info[-1]['line'])] *= alpha
            else:
                for j in line_num:
                    line_sus[int(j)] *= (2-alpha)
    # print(line_sus)
    if line_sus.count(min(line_sus)) >= file_length * beta:
        for i, line in enumerate(lines):
            flag = False
            if language == 'cpp' or language == 'c':
                flag = (util.find_pos('cin', line) or util.find_pos('printf', line))
            elif language == 'py':
                flag = util.find_pos('print', line)
            # print(line)
            # print(util.find_pos('printf', line))
            if flag == True:
                # print(line, i)
                line_sus[i + 1] *= alpha
    print(line_sus)
    return line_sus

def run_file(file_path, ac_file, test_dir_path, language):
    '''
    计算程序的最后怀疑度列表
    '''
    print(file_path, ac_file)
    line_sus = cal_sus(file_path, ac_file, test_dir_path, language)
    line_rank = []
    for i,sus in enumerate(line_sus):
        if i == 0:
            continue
        line_rank.append({
            'sus': sus,
            'no': i
        })
    line_rank.sort(key=lambda s: s['sus'])
    # print(line_rank)
    final_rank = []
    tmp_list = []
    for i in range(len(line_rank)):
        if i != 0 and abs(line_rank[i]['sus'] - line_rank[i-1]['sus']) > 0.0001:
            final_rank.append(tmp_list)
            tmp_list = []
        tmp_list.append(line_rank[i]['no'])
    if len(tmp_list) != 0:
        final_rank.append(tmp_list)
    print(final_rank)
    return final_rank, line_sus

def run_dir(file_dir_path, pair_info, test_dir_path):
    '''
    计算某个文件夹内所有文件的最后怀疑度列表
    '''
    file_list = os.listdir(file_dir_path)
    wb = openpyxl.load_workbook(res_file)
    if sys.platform == "linux":
        problem_id = file_dir_path.split('/')[-2]
    else:
        problem_id = file_dir_path.split('\\')[-2]
    wb.create_sheet(problem_id)
    ws = wb[problem_id]
    ws.append({'a':'name', 'b':'suspicion_rank', 'c':'suspicion'})
    for file in file_list:
        file_type = file.split('.')[-1]
        if file_type == 'c' or file_type == 'cpp' or file_type == 'py':
            wa_file_path = os.path.join(file_dir_path, file)
            ac_file_path = os.path.join(data_path, str(problem_id), 'AC_'+file_type, pair_info[file])
            # print(wa_file_path, ac_file_path)
            try:
                final_rank, line_sus = run_file(wa_file_path, ac_file_path, test_dir_path, file_type)
                ws.append({'a':file, 'b':str(final_rank), 'c': str(line_sus)})
                # util.add_file(res_file, file + '    ' + str(final_rank) + '    ' + str(VSBFL_rank) + '\n')
            except Exception:
                print(Exception)
                ws.append({'a':file, 'b':'[]', 'c':'[]'})
                continue
                # util.add_file(res_file, file + '    ' + 'contains error\n')
        # break
    wb.save(res_file)
    return

if __name__ == "__main__":
    # wa_file = r'C:\Users\ShenJitao\Desktop\cppsnooper\instrumentation\test\WA_c\wa.c'
    # ac_file = r'C:\Users\ShenJitao\Desktop\cppsnooper\instrumentation\test\AC_c\ac.c'
    # test_dir_path = r'C:\Users\ShenJitao\Desktop\cppsnooper\instrumentation\test\TEST_DATA_TCG1'
    # wa_file = r'D:\fault_loc\ITSP-data\2871\WA_c\278496_buggy.c'
    # ac_file = r'D:\fault_loc\ITSP-data\2871\AC_c\278496_correct.c'
    # test_dir_path = r'D:\fault_loc\ITSP-data\2871\TEST_DATA_TCG1'
    # rank = run_file(wa_file, ac_file, test_dir_path, 'cpp')
    
    res_file = res_file_temp
    # problem_list = [2810, 2811, 2812, 2813, 2824, 2825, 2827, 2828, 2830, 2831, 2832, 2833, 2864, 2865, 2866, 2867, 2868, 2869, 2870, 2871]      
    problem_list = ['2174', '3310', '3904', '3905', '3919', '3922', '3924']

    for problem in problem_list:
         # problem = 2810
        print(problem)
        pair_info = find_pair_by_tag(os.path.join(data_path, str(problem), 'Tag_c'))
        # pair_info = find_pair_by_res(r'D:\fault_loc\VSFL-TCG\result\cluster_op2.xlsx')
        dir_path = os.path.join(data_path, str(problem), 'WA_c')
        test_dir_path = os.path.join(data_path, str(problem), 'TEST_DATA_TCG1')
        run_dir(dir_path, pair_info, test_dir_path)
        # cal_time(dir_path)
    #     break

'''
[1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 2.0736, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0, 1.0]
[[1, 2, 3, 4, 5, 6, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21], [7]]
'''
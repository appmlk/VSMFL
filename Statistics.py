# -*- coding: utf-8 -*-
import openpyxl
import os
import util
import numpy as np

def get_all_tag_info(root_path, type='c'):
    '''
    获取数据集中所有真实的错误信息
    type选择何种类型语言（当然数据集里面只有c）
    '''
    problem_list = ['2810', '2811', '2812', '2813', '2824', '2825', '2827', '2828',
     '2830', '2831', '2832', '2833', '2864', '2865', '2866', '2867', '2868', '2869', '2870', '2871']      
    
    tag_info = {}
    dirs = os.listdir(root_path)
    for dir in dirs:
        if dir not in problem_list:
            continue
        # print(dir)
        tag_dir = os.path.join(root_path, dir, 'Tag_' + type)
        file_list = os.listdir(tag_dir)
        for i in file_list:
            file_path = os.path.join(tag_dir, i)
            lines = util.read_file(file_path)
            name = lines[0].replace('\n', '')
            arr = lines[2].split(',')
            tag_info[name] = []
            for num in arr:
                tag_info[name].append(int(num))
    # print(tag_info)
    return tag_info

def get_tag_info(tag_dir):
    '''
    获取某个题目下真实的错误信息
    '''
    tag_info = {}
    file_list = os.listdir(tag_dir)
    for i in file_list:
        file_path = os.path.join(tag_dir, i)
        lines = util.read_file(file_path)
        name = lines[0].replace('\n', '')
        arr = lines[2].split(',')
        tag_info[name] = []
        for num in arr:
            tag_info[name].append(int(num))
    return tag_info


def cal_top_N_multi(rows, tag_info):
    '''
    统计Top-N的值(多错误)
    '''
    top_1 = 0
    top_3 = 0
    top_5 = 0
    top_10 = 0
    cnt = 0
    for i, row in enumerate(rows):
        # print(row[0].value, end='\t')
        # print(row[1].value, end='\t')
        # print(row[2].value)
        if i == 0:
            continue
        cnt += 1
        name = row[0].value
        final_rank = eval(row[1].value)
        if len(final_rank) == 0:
            continue
        ans = 0
        for j, column in enumerate(final_rank):
            ans += len(column)
            for num in tag_info[name]:
                if num in column:
                    if ans <= 1:
                        top_1 += 1
                    elif ans <= 3:
                        top_3 += 1
                    elif ans <= 5:
                        top_5 += 1
                    elif ans <= 10:
                        top_10 += 1
        # break
    # return top_1/cnt, top_3/cnt, top_5/cnt, top_10/cnt
    return top_1, top_3, top_5, top_10

def cal_top_N_first(rows, tag_info):
    '''
    统计Top-N的值(第一个错误)
    '''
    top_1 = 0
    top_3 = 0
    top_5 = 0
    top_10 = 0
    cnt = 0
    for i, row in enumerate(rows):
        if i == 0:
            continue
        # print(row[0].value, end='\t')
        # print(row[1].value, end='\t')
        # print(row[2].value)
        cnt += 1
        name = row[0].value
        final_rank = eval(row[1].value)
        if len(final_rank) == 0:
            continue
        ans = 0
        for j, column in enumerate(final_rank):
            flag = 0
            for num in tag_info[name]:
                ans += len(column)
                if num in column:
                    flag = 1
                    break
            if flag == 1:
                break
        if ans <= 1:
            top_1 += 1
        elif ans <= 3:
            top_3 += 1
        elif ans <= 5:
            top_5 += 1
        elif ans <= 10:
            top_10 += 1
        # break
    # return top_1/cnt, top_3/cnt, top_5/cnt, top_10/cnt
    return top_1, top_3, top_5, top_10

def cal_exam_multi(rows, tag_info):
    '''
    计算exam值(多错误)
    '''
    res = []
    for i, row in enumerate(rows):
        # print(row[0].value, end='\t')
        # print(row[1].value, end='\t')
        # print(row[2].value)
        if i == 0:
            continue
        ans = 0
        name = row[0].value
        final_rank = eval(row[1].value)
        if len(final_rank) == 0:
            continue
        for num in tag_info[name]:
            no = 0
            for j, column in enumerate(final_rank):
                no += len(column)
                if num in column:
                    ans += no
                    break
        ans = ans / len(tag_info[name])
        res.append(ans)
    return res

def cal_exam_first(rows, tag_info):
    '''
    计算exam值(第一个错误)
    '''
    res = []
    for i, row in enumerate(rows):
        if i == 0:
            continue
        # print(row[0].value, end='\t')
        # print(row[1].value, end='\t')
        # print(row[2].value)
        name = row[0].value
        final_rank = eval(row[1].value)
        if len(final_rank) == 0:
            continue
        ans = 0
        for j, column in enumerate(final_rank):
            ans += len(column)
            flag = 0
            for num in tag_info[name]:
                if num in column:
                    flag = 1
                    break
            if flag == 1:
                break
        # if flag == 0:
        #     ans = 
        res.append(ans)
    return res

def get_code_exam(sheet, tag_info):
    '''
    '''
    sum = 0
    for i, row in enumerate(sheet.rows):
        if i == 0:
            continue
        name = row[0].value
        final_rank = eval(row[1].value)
        if len(final_rank) == 0:
            continue
        res1 = 0
        ans1 = 0
        res2 = 0
        ans2 = 1
        flag = 1
        for j, column in enumerate(final_rank):
            ans1 += len(column)
            for num in tag_info[name]:
                if num in column:
                    # if flag == 1:
                        # res1 += ans1
                        # res2 += ans2
                        # flag = 0
                    res1 += ans1
                    res2 += ans2
                    flag = 0
            ans2 += len(column)
        if flag == 1:
            res1 = ans1
            res2 = ans1
        else:
            res1 = res1 / len(tag_info[name])
            res2 = res2 / len(tag_info[name])
        if ans1 != 0:
            sheet.cell(i + 1, 4).value = (res1 + res2) / 2 / ans1
            # sheet.cell(i + 1, 5).value = (res1 + res2) / 2 / ans1
            # sheet.cell(i + 1, 6).value = ''
        else:
            sheet.cell(i + 1, 4).value = 1
            # sheet.cell(i + 1, 5).value = 1
            # sheet.cell(i + 1, 6).value = ''
        sum += (res1 + res2) / 2 / ans1
    return sum

def statistical_fl_results(file_path, tag_root_dir):
    '''
    统计错误定位的实验结果
    '''
    if file_path.split('.')[-1] != 'xlsx' and file_path.split('.')[-1] != 'xls':
        print('not an execl file')
        return
    wb = openpyxl.load_workbook(file_path) # 读取 工作簿
    ws_first = wb.worksheets[0]  
    ws_first.title = 'first'
    ws_multi = wb.worksheets[1]
    ws_multi.title = 'multi'
    ws_first.append({'a':'problem_id', 'b':'exam', 'c':'top_1', 'd':'top_3', 'e':'top_5', 'f': 'top_10','g': 'total_num'})
    ws_multi.append({'a':'problem_id', 'b':'exam', 'c':'top_1', 'd':'top_3', 'e':'top_5', 'f': 'top_10','g': 'total_num'})
    ans1 = 0
    ans2 = 0
    for i, sheet in enumerate(wb):
        if i == 0 or i == 1:
            continue
        # if sheet.title != '2812':
        #     continue
        print(sheet.title)
        # row_max = sheet.max_row # 获取最大行
        # con_max = sheet.max_column # 获取最大列
        tag_dir = os.path.join(tag_root_dir, sheet.title, 'Tag_c')
        tag_info = get_tag_info(tag_dir)
        # for i in tag_info:
        #     ans1 += len(tag_info[i])
        # continue
        ans1 += get_code_exam(sheet, tag_info) # 给每道题添加exam
        ans2 += (sheet.max_row - 1)
        exam = cal_exam_first(sheet.rows, tag_info)
        top_1, top_3, top_5, top_10 = cal_top_N_first(sheet.rows, tag_info)
        ws_first.append({'a':sheet.title, 'b':np.mean(exam), 'c':top_1, 'd':top_3, 'e':top_5, 'f': top_10, 'g':sheet.max_row - 1})
        
        top_1, top_3, top_5, top_10 = cal_top_N_multi(sheet.rows, tag_info)
        exam = cal_exam_multi(sheet.rows, tag_info)
        ws_multi.append({'a':sheet.title, 'b':np.mean(exam), 'c':top_1, 'd':top_3, 'e':top_5, 'f': top_10, 'g':sheet.max_row - 1})
        
        # print(exam)
        # print(top_1, top_3, top_5, top_10)
        # break
    print(ans1, ans2, ans1/ans2)
    wb.save(file_path)
    return

def cal_exact_result(rows):
    '''
    计算匹配到的代码包含数据集配对代码的数量
    '''
    name = ''
    cnt = 0   #代码份数
    ans1 = 0  #统计结果包含配对代码的
    ans2 = 0  #统计结果恰好是配对代码的
    for i, row in enumerate(rows):
        if i == 0:
            continue
        if row[0].value != None:
            cnt += 1
            name = row[0].value.split('_')[0]
        if row[1].value.find('[') >= 0:
            res_list = eval(row[1].value)
            res_list = list(map(lambda x: x.split('_')[0], res_list))
            # print(res_list)
            if name in res_list:
                ans1 += 1
            if len(res_list) == 1 and res_list[0] == name:
                ans2 += 1
    print(ans1, ans2, cnt)
    return ans1, ans2, ans1/cnt, ans2/cnt

def statistical_parse_results(file_path):
    '''
    统计匹配正确代码的结果
    '''
    if file_path.split('.')[-1] != 'xlsx' and file_path.split('.')[-1] != 'xls':
        print('not an execl file')
        return
    wb = openpyxl.load_workbook(file_path)
    ws = wb.worksheets[0]
    ws.append({'a':'name', 'b':'contain_cnt', 'c':'contain_ratio', 'd':'exact_cnt', 'e':'exact_ratio'})
    for i, sheet in enumerate(wb):
        if i == 0:
            continue
        print(sheet.title)
        ans1, ans2, ratio1, ratio2 = cal_exact_result(sheet.rows)
        ws.append({'a':sheet.title, 'b':ans1, 'c':ratio1, 'd': ans2, 'e': ratio2})
        # break
    wb.save(file_path)
    return

def statistical_type_result(file_path, tag_root_dir):
    '''
    '''
    if file_path.split('.')[-1] != 'xlsx' and file_path.split('.')[-1] != 'xls':
        print('not an execl file')
        return
    wb = openpyxl.load_workbook(file_path) # 读取 工作簿
    ws = wb.worksheets[0]
    ws.append({'a': 'type', 'b': 'count', 'c': 'first', 'd': 'multi'})
    simple_expression = {
        'count': 0, 'first': 0, 'multi': 0,
    }
    conditions = {
        'first': 0, 'multi': 0, 'count': 0,
    }
    loops = {
        'first': 0, 'multi': 0, 'count': 0,
    }
    for i, sheet in enumerate(wb):
        if i == 0:
            continue
        print(sheet.title)
        for j, row in enumerate(sheet.rows):
            if j == 0:
                continue
            if sheet.title in ['2810', '2811', '2812', '2813']:
                # simple_expression['first'] += row[4].value
                simple_expression['multi'] += row[3].value
                simple_expression['count'] += 1
            elif sheet.title in ['2824', '2825', '2827', '2828', '2830', '2831', '2832', '2833']:
                # conditions['first'] += row[4].value
                conditions['multi'] += row[3].value
                conditions['count'] += 1
            elif sheet.title in ['2864', '2865', '2866', '2867', '2868', '2869', '2870', '2871']:
                # loops['first'] += row[4].value
                loops['multi'] += row[3].value
                loops['count'] += 1
    print(simple_expression, conditions, loops)
    ws.append({'a': 'simple_expression', 'b': simple_expression['count'], 'c': simple_expression['first'] / simple_expression['count'], 'd': simple_expression['multi'] / simple_expression['count']})
    ws.append({'a': 'conditions', 'b': conditions['count'], 'c': conditions['first'] / conditions['count'], 'd': conditions['multi'] / conditions['count']})
    ws.append({'a': 'loops', 'b': loops['count'], 'c': loops['first'] / loops['count'], 'd': loops['multi'] / loops['count']})
    wb.save(file_path)
    return

def cal_blank_line(file_path, file_path2):
    '''
    查看excel表中有多少空行
    空行代表执行失败的代码
    '''
    if file_path.split('.')[-1] != 'xlsx' and file_path.split('.')[-1] != 'xls':
        print('not an execl file')
        return
    del_list = ['363003.c', '192833.c', '494623.c', '357210.c', '495301.c', '190500.c', '197305.c', '287729.c', '199201.c', '357216.c', '516537.c', '198800.c', '159249.c', '359662.c', '523972.c', '158224.c', '358221.c', '197195.c', '484121.c', '356712.c', '198148.c', '523958.c', '246723.c', '495291.c', '435518.c', '83830.c', '59289.c', '287668.c', '357025.c', '365254.c', '443625.c', '151928.c', '287550.c', '198167.c', '158695.c', '249663.c', '129137.c', '523618.c', '288163.c', '359652.c', '523951.c', '287704.c', '306598.c', '287843.c', '210215.c', '491151.c', '198804.c', '210364.c', '358189.c', '523617.c', '540166.c', '198317.c', '435260.c', '198334.c', '357221.c', '484000.c', '198160.c', '158293.c', '435291.c', '546063.c', '440774.c', '306590.c', '493541.c', '199058.c', '491149.c', '363006.c', '357227.c', '154862.c', '494439.c', '288549.c', '356228.c', '363398.c', '160804.c', '245230.c', '199205.c', '358603.c', '546011.c', '546039.c', '158297.c', '246722.c', '484002.c', '435244.c', '198327.c', '84856.c', '438546.c', '438455.c', '158254.c', '158682.c', '158300.c', '287542.c', '523849.c', '151927.c', '156381.c', '357215.c', '435218.c', '500041.c', '363017.c', '528772.c', '356915.c', '86017.c', '151929.c', '546030.c', '194721.c', '528291.c', '520344.c', '250210.c', '363005.c', '198335.c', '198264.c', '246354.c', '199070.c', '158878.c', '198152.c', '520349.c', '158696.c', '155224.c', '198330.c', '484118.c', '491544.c', '495299.c', '356229.c', '435463.c', '160626.c', '151967.c', '483994.c', '198563.c']
    wb = openpyxl.load_workbook(file_path) # 读取工作簿
    new_wb = openpyxl.load_workbook(file_path2)
    for i, sheet in enumerate(wb):
        if i == 0:
            continue
        print(sheet.title)
        new_wb.create_sheet(sheet.title)
        new_ws = new_wb[sheet.title]
        # new_ws.append({'a':'name', 'b':'suspicion_rank', 'c':'suspicion'})
        for j, row in enumerate(sheet):
            name = row[0].value
            if name not in del_list:
                # print(name)
                new_ws.append({'a': name, 'b': row[1].value, 'c':row[2].value})
    new_wb.save(file_path2)
    return

if __name__ == "__main__":

    # file_path = r'C:\Users\ShenJitao\Desktop\cppsnooper\instrumentation\result\sbfl(Jaccard)_sus.xlsx'
    file_path = r'D:\fault_loc\VSFL-TCG\result\fault_loc_sbfl(Jaccard).xlsx'
    tag_dir = r'D:\fault_loc\ITSP-data'
    statistical_fl_results(file_path, tag_dir)
    # statistical_type_result(file_path, tag_dir)

    # file_path = r'C:\Users\ShenJitao\Desktop\cppsnooper\instrumentation\result\result\vsbfl_BUCT_sus.xlsx'
    # new_file_path = r'C:\Users\ShenJitao\Desktop\cppsnooper\instrumentation\result\vsbfl_BUCT_sus.xlsx'
    # cal_blank_line(file_path, new_file_path)
    # for fm in ['Tarantula', 'Dstar', 'Ochiai', 'Op2']:
    #     file_path = file_tmp_path % fm
    #     new_file_path = new_file_tmp_path % fm
    #     cal_blank_line(file_path, new_file_path)
        # break
    # file_path = r'C:\Users\ShenJitao\Desktop\cppsnooper\instrumentation\result\vsbfl_BUCT_sus.xlsx'
    # print(cal_blank_line(file_path))
    # file_path = r'E:\fault_loc\VSFL-TCG\result\cluster_op2.xlsx'
    # statistical_parse_results(file_path)